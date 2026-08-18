from __future__ import annotations

import csv
import queue
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from controller import AppController, _QueueDrainSignal, _is_proxy_network_error
from models import QUERY_BACKEND_HTTP, ProfileResult, RunConfig
from proxy_pool import ProxySpec


class FakeBrowserSession:
    observed_input_counts: list[int] = []

    def __init__(self, **_kwargs: object) -> None:
        self.page = None

    def start(self, fresh_profile: bool = False) -> None:
        self.page = object()

    def process(self, person: object) -> list[ProfileResult]:
        source = getattr(person, "source_path")
        if source.suffix.lower() in {".xlsx", ".xlsm"}:
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
            try:
                self.observed_input_counts.append(max(0, workbook.active.max_row - 1))
            finally:
                workbook.close()
        else:
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                self.observed_input_counts.append(max(0, len(list(csv.reader(handle))) - 1))
        return [
            ProfileResult(1, "https://www.mylife.com/jane-doe/e1", full_name="Jane Doe", birthday="March 7, 1984", gender="Female", zodiac="Pisces (February 19 - March 20)"),
        ]

    def clear_person_data(self) -> None:
        pass

    def close(self) -> None:
        pass

    def rebuild(self, proxy_geo: object = None) -> None:
        self.page = object()


class StoppingFakeBrowserSession(FakeBrowserSession):
    def __init__(self, **kwargs: object) -> None:
        super().__init__(**kwargs)
        self.stop_event = kwargs["stop_event"]

    def process(self, person: object) -> list[ProfileResult]:
        results = super().process(person)
        self.stop_event.set()
        return results


class WrappedCancellationFakeBrowserSession(FakeBrowserSession):
    """模拟底层把停止包装成 RuntimeError，而不是直接抛出 Cancelled。"""

    def __init__(self, **kwargs: object) -> None:
        super().__init__(**kwargs)
        self.stop_event = kwargs["stop_event"]

    def process(self, person: object) -> list[ProfileResult]:
        self.stop_event.set()
        raise RuntimeError("cancelled")

    def capture_diagnostic(self, reason: str) -> None:
        raise AssertionError(f"停止期间不应继续截图：{reason}")


class PersistentWorkerFakeBrowserSession(FakeBrowserSession):
    """让一个任务阻塞，验证暂时空队列不会让另一个 worker 提前退出。"""

    slow_started = threading.Event()
    fast_done = threading.Event()
    release_slow = threading.Event()

    def process(self, person: object) -> list[ProfileResult]:
        if getattr(person, "first_name", "") == "Jane":
            self.slow_started.set()
            if not self.release_slow.wait(10):
                raise TimeoutError("test did not release slow worker")
        else:
            self.fast_done.set()
        return [
            ProfileResult(
                1,
                f"https://www.mylife.com/{getattr(person, 'first_name', 'person').lower()}/e1",
                full_name=f"{getattr(person, 'first_name', '')} {getattr(person, 'last_name', '')}".strip(),
                birthday="March 7, 1984",
                gender="Female",
                zodiac="Pisces (February 19 - March 20)",
            )
        ]


class RepeatingConnectivityFakePool:
    """模拟代理一直不可用；任务结束信号到达后必须立即退出检查循环。"""

    def __init__(self) -> None:
        self.checks = 0

    def wait_until_ready(self, _spec: object, cancel: object) -> None:
        for _ in range(200):
            if cancel.is_set():
                return None
            self.checks += 1
            if cancel.wait(0.002):
                return None
        raise AssertionError("任务全部结束后仍在反复测试 IP 连通性")


class ControllerOfflineTests(unittest.TestCase):
    def test_http_backend_selects_http_interface_worker(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text("first_name,last_name\nJane,Doe\n", encoding="utf-8-sig")
            controller = AppController(
                RunConfig(
                    source,
                    root / "output",
                    thread_count=1,
                    browser_mode="无头",
                    query_backend=QUERY_BACKEND_HTTP,
                )
            )
            with patch("controller.HttpInterfaceSession", FakeBrowserSession), patch(
                "controller.BrowserSession"
            ) as browser_session:
                controller.run()
            browser_session.assert_not_called()
            self.assertEqual(controller._final_status, "已完成")

    def test_stop_does_not_deadlock_behind_background_gui_log_callback(self) -> None:
        """后台日志等待 Tk 主线程时，停止回调不能反向等待同一把日志锁。"""

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text("first_name,last_name\nJane,Doe\n", encoding="utf-8-sig")
            callback_entered = threading.Event()
            release_callback = threading.Event()

            def gui_log(_message: str) -> None:
                if threading.current_thread().name == "background-log-producer":
                    callback_entered.set()
                    release_callback.wait(5)

            controller = AppController(
                RunConfig(source, root / "output", thread_count=1, browser_mode="无头"),
                log=gui_log,
            )
            producer = threading.Thread(
                target=lambda: controller._log("后台日志"),
                name="background-log-producer",
            )
            stopper = threading.Thread(target=controller.stop, name="simulated-gui-thread")
            producer.start()
            try:
                self.assertTrue(callback_entered.wait(2), "后台日志回调未进入等待状态")
                stopper.start()
                self.assertTrue(controller.stop_event.wait(0.5), "停止信号被日志锁延迟")
                stopper.join(0.5)
                self.assertFalse(stopper.is_alive(), "GUI 停止回调被后台日志锁死")
            finally:
                release_callback.set()
                producer.join(2)
                stopper.join(2)

    def test_proxy_browser_connection_error_is_detected_for_refresh(self) -> None:
        self.assertTrue(_is_proxy_network_error("Page.goto: net::ERR_SOCKS_CONNECTION_FAILED"))
        self.assertTrue(_is_proxy_network_error("Page.goto: net::ERR_SSL_PROTOCOL_ERROR"))
        self.assertTrue(_is_proxy_network_error("Page.goto: net::ERR_TIMED_OUT"))
        self.assertFalse(_is_proxy_network_error("selector not found"))

    def test_proxy_recovery_wait_stops_when_database_has_no_unfinished_jobs(self) -> None:
        held_jobs: queue.Queue[tuple[int, object, int]] = queue.Queue()
        held_jobs.put((1, object(), 0))
        held_jobs.put((2, object(), 0))
        signal = _QueueDrainSignal(
            threading.Event(),
            held_jobs,
            finished_threshold=1,
            finished_check=lambda: True,
        )
        self.assertTrue(signal.is_set())

    def test_proxy_connectivity_wait_stops_when_all_jobs_are_finished(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text("first_name,last_name\n", encoding="utf-8-sig")
            logs: list[str] = []
            controller = AppController(
                RunConfig(source, root / "output", thread_count=1, browser_mode="无头"),
                log=logs.append,
            )
            spec = ProxySpec(1, "proxy.example", 1080, "user", "pass", "https://refresh.example/change")

            empty_queue: queue.Queue[tuple[int, object, int]] = queue.Queue()
            empty_pool = RepeatingConnectivityFakePool()
            controller._worker(1, empty_queue, object(), spec, empty_pool)
            self.assertEqual(empty_pool.checks, 0)

            inflight_queue: queue.Queue[tuple[int, object, int]] = queue.Queue()
            inflight_queue.put((1, object(), 0))
            waiting_pool = RepeatingConnectivityFakePool()

            def finish_elsewhere() -> None:
                time.sleep(0.02)
                inflight_queue.get_nowait()
                inflight_queue.task_done()

            finisher = threading.Thread(target=finish_elsewhere)
            finisher.start()
            controller._worker(1, inflight_queue, object(), spec, waiting_pool)
            finisher.join(1)

            self.assertFalse(finisher.is_alive())
            self.assertGreater(waiting_pool.checks, 0)
            self.assertTrue(any("任务已全部完成，停止连通性检测" in line for line in logs))

    def test_completed_input_finishes_without_starting_proxy_health_loop(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text("first_name,last_name\n", encoding="utf-8-sig")
            pool = RepeatingConnectivityFakePool()
            controller = AppController(
                RunConfig(
                    source,
                    root / "output",
                    thread_count=1,
                    browser_mode="无头",
                    proxy_enabled=True,
                    proxy_lines=["proxy.example:1080:user:pass|https://refresh.example/change"],
                )
            )

            with (
                patch("controller.ProxyPool", return_value=pool),
                patch("controller.BrowserSession") as browser,
            ):
                controller.run()

            self.assertEqual(controller._final_status, "已完成")
            self.assertEqual(pool.checks, 0)
            browser.assert_not_called()

    def test_full_pipeline_deletes_each_explicit_result_in_realtime(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text(
                "first_name,last_name,note\nJane,Doe,first\nJohn,Doe,second\n",
                encoding="utf-8-sig",
            )
            output = root / "output"
            config = RunConfig(source, output, thread_count=1, browser_mode="无头")
            progress_states: list[dict[str, int | str]] = []
            controller = AppController(config, progress=progress_states.append)
            FakeBrowserSession.observed_input_counts = []
            with patch("controller.BrowserSession", FakeBrowserSession):
                controller.run()
            current_run_states = [state for state in progress_states if state.get("total") == 2]
            self.assertTrue(current_run_states)
            self.assertTrue(any(state.get("output", "").endswith("people_MyLife结果.csv") for state in current_run_states))
            self.assertEqual(progress_states[-1].get("status"), "已完成")
            self.assertEqual(progress_states[-1].get("completed"), 2)
            self.assertEqual(progress_states[-1].get("pending"), 0)
            self.assertEqual(progress_states[-1].get("failed"), 0)
            result_path = output / "people_MyLife结果.csv"
            with result_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 2)
            self.assertEqual([row["note"] for row in rows], ["first", "second"])
            self.assertEqual(list(rows[0])[-4:], ["生日", "性别", "星座", "备注原因"])
            self.assertIn("搜索方式：", rows[0]["备注原因"])
            self.assertEqual(rows[0]["生日"], "March 7, 1984")
            self.assertEqual(rows[0]["性别"], "Female")
            self.assertEqual(rows[0]["星座"], "Pisces (February 19 - March 20)")
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                remaining = list(csv.reader(handle))
            self.assertEqual(remaining, [["first_name", "last_name", "note"]])
            self.assertEqual(sorted(FakeBrowserSession.observed_input_counts), [1, 2])
            self.assertFalse((output / ".mylife_runtime" / "profiles").exists())

    def test_xlsx_keeps_all_rows_during_processing_then_deletes_once(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.append(["first_name", "last_name", "note"])
            workbook.active.append(["Jane", "Doe", "first"])
            workbook.active.append(["John", "Doe", "second"])
            workbook.create_sheet("keep").append(["other sheet", "must stay"])
            workbook.save(source)
            workbook.close()
            controller = AppController(RunConfig(source, root / "output", thread_count=1, browser_mode="无头"))
            FakeBrowserSession.observed_input_counts = []
            with patch("controller.BrowserSession", FakeBrowserSession):
                controller.run()
            self.assertEqual(FakeBrowserSession.observed_input_counts, [2, 2])
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
            try:
                self.assertEqual(list(workbook.active.iter_rows(values_only=True)), [("first_name", "last_name", "note")])
                self.assertEqual(list(workbook["keep"].iter_rows(values_only=True)), [("other sheet", "must stay")])
            finally:
                workbook.close()

    def test_stopped_xlsx_waits_for_worker_then_bulk_deletes_completed_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.append(["first_name", "last_name", "note"])
            workbook.active.append(["Jane", "Doe", "done"])
            workbook.active.append(["John", "Doe", "pending"])
            workbook.create_sheet("keep").append(["other sheet", "must stay"])
            workbook.save(source)
            workbook.close()
            events: list[str] = []
            controller = AppController(
                RunConfig(source, root / "output", thread_count=1, browser_mode="无头"),
                progress=lambda state: events.append(str(state.get("status", ""))),
            )
            real_cleanup = controller._cleanup_cache

            def observed_cleanup() -> None:
                real_cleanup()
                events.append("cleanup")

            controller._cleanup_cache = observed_cleanup
            StoppingFakeBrowserSession.observed_input_counts = []
            with patch("controller.BrowserSession", StoppingFakeBrowserSession):
                controller.run()
            self.assertEqual(StoppingFakeBrowserSession.observed_input_counts, [2])
            self.assertEqual(controller._final_status, "已停止")
            self.assertEqual(events[-2:], ["cleanup", "已停止"])
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    list(workbook.active.iter_rows(values_only=True)),
                    [("first_name", "last_name", "note"), ("John", "Doe", "pending")],
                )
                self.assertEqual(list(workbook["keep"].iter_rows(values_only=True)), [("other sheet", "must stay")])
            finally:
                workbook.close()

    def test_wrapped_cancellation_during_stop_is_retry_not_failed(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.append(["first_name", "last_name", "note"])
            workbook.active.append(["Jane", "Doe", "must retry"])
            workbook.save(source)
            workbook.close()

            logs: list[str] = []
            controller = AppController(
                RunConfig(source, root / "output", thread_count=1, browser_mode="无头"),
                log=logs.append,
            )
            with patch("controller.BrowserSession", WrappedCancellationFakeBrowserSession):
                controller.run()

            summary = controller.database.summary(source)
            self.assertEqual(controller._final_status, "已停止")
            self.assertEqual(summary.get("retry"), 1)
            self.assertEqual(summary.get("failed", 0), 0)
            connection = controller.database.connect()
            try:
                attempts = connection.execute("SELECT attempts FROM jobs").fetchone()[0]
            finally:
                connection.close()
            self.assertEqual(attempts, 0, "用户停止不应消耗人员任务的业务重试次数")
            self.assertTrue(any("任务已保留为可续跑状态" in line for line in logs))
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    list(workbook.active.iter_rows(values_only=True)),
                    [("first_name", "last_name", "note"), ("Jane", "Doe", "must retry")],
                )
            finally:
                workbook.close()

    def test_proxy_thread_limit_validation(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.txt"
            source.write_text("Jane Doe\n", encoding="utf-8")
            config = RunConfig(
                source,
                root / "out",
                thread_count=2,
                proxy_enabled=True,
                proxy_lines=["proxy.example:1080:u:p|https://refresh.example/change"],
            )
            controller = AppController(config)
            with self.assertRaisesRegex(ValueError, "线程数量不可以超过代理数量"):
                controller._validate()

    def test_new_run_repairs_legacy_exhausted_retry_row(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text(
                "first_name,last_name,note\nJane,Doe,legacy stopped row\n",
                encoding="utf-8-sig",
            )
            logs: list[str] = []
            controller = AppController(
                RunConfig(source, root / "output", thread_count=1, browser_mode="无头"),
                log=logs.append,
            )
            from input_loader import load_people

            controller.database.import_people(load_people(source)[1])
            connection = controller.database.connect()
            try:
                connection.execute(
                    "UPDATE jobs SET status='retry', attempts=?",
                    (controller.config.max_job_attempts,),
                )
                connection.commit()
            finally:
                connection.close()
            with patch("controller.BrowserSession", FakeBrowserSession):
                controller.run()
            self.assertEqual(controller._final_status, "已完成")
            self.assertTrue(any("旧停止流程遗留的不可领取行" in line for line in logs))
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                self.assertEqual(list(csv.reader(handle)), [["first_name", "last_name", "note"]])

    def test_workers_wait_for_inflight_tasks_instead_of_disappearing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "people.csv"
            source.write_text(
                "first_name,last_name,note\nJane,Doe,slow\nJohn,Doe,fast\n",
                encoding="utf-8-sig",
            )
            controller = AppController(
                RunConfig(source, root / "output", thread_count=2, browser_mode="无头")
            )
            PersistentWorkerFakeBrowserSession.slow_started = threading.Event()
            PersistentWorkerFakeBrowserSession.fast_done = threading.Event()
            PersistentWorkerFakeBrowserSession.release_slow = threading.Event()
            runner = threading.Thread(target=controller.run, name="controller-test-runner")
            with patch("controller.BrowserSession", PersistentWorkerFakeBrowserSession):
                runner.start()
                try:
                    self.assertTrue(PersistentWorkerFakeBrowserSession.slow_started.wait(5))
                    self.assertTrue(PersistentWorkerFakeBrowserSession.fast_done.wait(5))
                    time.sleep(1.1)
                    self.assertEqual(
                        sum(thread.is_alive() for thread in controller.worker_threads),
                        2,
                        "队列暂时为空时两个 worker 都应等待未完成任务，不能提前丢线程",
                    )
                finally:
                    PersistentWorkerFakeBrowserSession.release_slow.set()
                    runner.join(15)
            self.assertFalse(runner.is_alive())
            self.assertEqual(controller._final_status, "已完成")


if __name__ == "__main__":
    unittest.main()
