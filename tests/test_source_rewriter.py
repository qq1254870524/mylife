from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

import openpyxl

from input_loader import load_people
from source_rewriter import RealtimeInputRewriter, remove_completed_rows


class SourceRewriterTests(unittest.TestCase):
    def test_csv_is_rebuilt_by_source_row_not_duplicate_first_value(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.csv"
            path.write_text("name,note\nJane Doe,a\nJohn Doe,b\nJane Doe,c\n", encoding="utf-8-sig")
            removed = remove_completed_rows(path, {2})
            self.assertEqual(removed, 1)
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows, [["name", "note"], ["John Doe", "b"], ["Jane Doe", "c"]])

    def test_xlsx_rebuild_preserves_unprocessed_rows(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["name", "note"])
            sheet.append(["Jane Doe", "a"])
            sheet.append(["John Doe", "b"])
            second = workbook.create_sheet("keep")
            second.append(["name", "note"])
            second.append(["Other Person", "must stay"])
            workbook.save(path)
            workbook.close()
            self.assertEqual(remove_completed_rows(path, {2}), 1)
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            other_rows = list(workbook["keep"].iter_rows(values_only=True))
            workbook.close()
            self.assertEqual(rows, [("name", "note"), ("John Doe", "b")])
            self.assertEqual(other_rows, [("name", "note"), ("Other Person", "must stay")])

    def test_realtime_csv_matches_full_row_and_removes_all_exact_duplicates(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.csv"
            path.write_text(
                "name,note\nJane Doe,a\nJane Doe,b\nJane Doe,a\n",
                encoding="utf-8-sig",
            )
            _headers, people = load_people(path)
            rewriter = RealtimeInputRewriter(path)
            self.assertEqual(rewriter.remove_person(people[0]), 2)
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows, [["name", "note"], ["Jane Doe", "b"]])
            self.assertEqual(rewriter.remove_person(people[1]), 1)
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows, [["name", "note"]])

    def test_realtime_xlsx_matches_original_and_preserves_other_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.append(["name", "note"])
            workbook.active.append(["Jane Doe", "delete"])
            workbook.active.append(["John Doe", "keep"])
            workbook.create_sheet("keep").append(["must", "stay"])
            workbook.save(path)
            workbook.close()
            _headers, people = load_people(path)
            self.assertEqual(RealtimeInputRewriter(path).remove_person(people[0]), 1)
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            self.assertEqual(
                list(workbook.active.iter_rows(values_only=True)),
                [("name", "note"), ("John Doe", "keep")],
            )
            self.assertEqual(list(workbook["keep"].iter_rows(values_only=True)), [("must", "stay")])
            workbook.close()

    def test_realtime_delimited_txt_keeps_header(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.txt"
            path.write_text("name,city\nJane Doe,Austin\nJohn Doe,Dallas\n", encoding="utf-8")
            _headers, people = load_people(path)
            self.assertEqual(RealtimeInputRewriter(path).remove_person(people[0]), 1)
            self.assertEqual(path.read_text(encoding="utf-8-sig"), "name,city\nJohn Doe,Dallas\n")


if __name__ == "__main__":
    unittest.main()
